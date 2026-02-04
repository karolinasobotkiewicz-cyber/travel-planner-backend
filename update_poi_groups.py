import openpyxl

# Otwórz Excel
wb = openpyxl.load_workbook('data/zakopane.xlsx')
ws = wb.active

# Znajdź kolumnę "Name" i "Target group"
header_row = 1
name_col = None
target_group_col = None

for col_idx, cell in enumerate(ws[header_row], start=1):
    if cell.value == "Name":
        name_col = col_idx
    elif cell.value == "Target group":
        target_group_col = col_idx

print(f"Name column: {name_col}, Target group column: {target_group_col}")

# Edytuj POI według wytycznych klientki
changes_made = []

for row_idx in range(2, ws.max_row + 1):
    name_cell = ws.cell(row=row_idx, column=name_col)
    target_group_cell = ws.cell(row=row_idx, column=target_group_col)
    
    name = name_cell.value
    current_groups = target_group_cell.value
    
    if name == "Park Harnasia":
        # Tylko family_kids
        new_groups = "family_kids"
        target_group_cell.value = new_groups
        changes_made.append(f"✓ {name}: '{current_groups}' → '{new_groups}'")
        print(f"ZMIANA: {name} | {current_groups} → {new_groups}")
    
    elif name == "Myszogród":
        # Usuń solo
        if current_groups:
            groups = [g.strip() for g in current_groups.split(',')]
            if 'solo' in groups:
                groups.remove('solo')
            new_groups = ', '.join(groups)
            target_group_cell.value = new_groups
            changes_made.append(f"✓ {name}: '{current_groups}' → '{new_groups}'")
            print(f"ZMIANA: {name} | {current_groups} → {new_groups}")
    
    elif name == "Dom do góry nogami":
        # Usuń solo
        if current_groups:
            groups = [g.strip() for g in current_groups.split(',')]
            if 'solo' in groups:
                groups.remove('solo')
            new_groups = ', '.join(groups)
            target_group_cell.value = new_groups
            changes_made.append(f"✓ {name}: '{current_groups}' → '{new_groups}'")
            print(f"ZMIANA: {name} | {current_groups} → {new_groups}")

# Zapisz Excel
wb.save('data/zakopane.xlsx')
print(f"\n✅ Zapisano {len(changes_made)} zmian:")
for change in changes_made:
    print(change)
