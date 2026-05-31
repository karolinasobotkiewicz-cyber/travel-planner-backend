"""
Remove 'Toporowa Cyrhla' from both data Excel files.
Backups are created before modification.
"""
import shutil
import pandas as pd
import openpyxl
from datetime import datetime

TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")

# ─── 1. zakopane.xlsx ────────────────────────────────────────────────────────
ZAKOPANE = 'data/zakopane.xlsx'
ZAKOPANE_BACKUP = f'data/zakopane_backup_{TIMESTAMP}.xlsx'
shutil.copy2(ZAKOPANE, ZAKOPANE_BACKUP)
print(f"Backup: {ZAKOPANE_BACKUP}")

df_z = pd.read_excel(ZAKOPANE)
before_z = len(df_z)
df_z = df_z[~df_z['Name'].str.contains('Toporowa', case=False, na=False)]
after_z = len(df_z)
df_z.to_excel(ZAKOPANE, index=False)
print(f"zakopane.xlsx: removed {before_z - after_z} row(s)  ({before_z} -> {after_z})")

# ─── 2. multi_city_attractions.xlsx (sheet: All Cities) ──────────────────────
MULTI = 'data/multi_city_attractions.xlsx'
MULTI_BACKUP = f'data/multi_city_backup_{TIMESTAMP}.xlsx'
shutil.copy2(MULTI, MULTI_BACKUP)
print(f"Backup: {MULTI_BACKUP}")

wb = openpyxl.load_workbook(MULTI)
ws = wb['All Cities']

# Find 'Name' column index (1-based)
header_row = [cell.value for cell in ws[1]]
try:
    name_col = header_row.index('Name') + 1  # 1-based
except ValueError:
    name_col = header_row.index('name') + 1

rows_to_delete = []
for row in ws.iter_rows(min_row=2):
    cell_value = row[name_col - 1].value
    if cell_value and 'Toporowa' in str(cell_value):
        rows_to_delete.append(row[0].row)

print(f"multi_city_attractions.xlsx: found {len(rows_to_delete)} row(s) to delete: {rows_to_delete}")
# Delete in reverse order to preserve row indices
for row_idx in reversed(rows_to_delete):
    ws.delete_rows(row_idx)

wb.save(MULTI)
print(f"multi_city_attractions.xlsx: saved ({len(rows_to_delete)} row(s) removed)")
print("\nDone.")
