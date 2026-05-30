import sys, os
sys.path.insert(0, '.')
os.environ['PYTHONIOENCODING'] = 'utf-8'

from app.infrastructure.repositories.load_multi_city import load_multi_city_poi
import openpyxl

# Test loading Trojmiasto cities
cities_troj = ['Gdańsk', 'Gdynia', 'Sopot']
pois = load_multi_city_poi('data/multi_city_attractions.xlsx', cities_troj)
print(f'Trojmiasto (Gdańsk/Gdynia/Sopot): {len(pois)} POIs')
if pois:
    p = pois[0]
    print(f'  Sample: {p.name} | tags: {p.tags[:5] if p.tags else []}')

# Kotlina
cities_kl = ['Kłodzko', 'Polanica-Zdrój', 'Kudowa-Zdrój']
pois3 = load_multi_city_poi('data/multi_city_attractions.xlsx', cities_kl)
print(f'Kotlina Kłodzka: {len(pois3)} POIs')
if pois3:
    p = pois3[0]
    print(f'  Sample: {p.name} | tags: {p.tags[:5] if p.tags else []}')
    print(f'  city: {p.city}')

# Check all unique cities in Excel
wb = openpyxl.load_workbook('data/multi_city_attractions.xlsx', read_only=True)
ws = wb.active
headers_row = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
city_col = headers_row.index('City') if 'City' in headers_row else None
if city_col is not None:
    cities = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[city_col]:
            cities.add(str(row[city_col]).strip())
    print(f'\nAll unique cities in multi_city_attractions.xlsx:')
    for c in sorted(cities):
        print(f'  {repr(c)}')
wb.close()

pass

tests = [
    ("Krakow (ascii)", {
        'location': {'city': 'Krakow', 'country': 'Poland', 'region_type': 'city'},
        'group': {'type': 'couples', 'size': 2, 'crowd_tolerance': 2},
        'trip_length': {'days': 2, 'start_date': '2026-06-07'},
        'daily_time_window': {'start': '09:00', 'end': '19:00'},
        'budget': {'level': 2},
        'transport_modes': ['walk'],
        'preferences': ['museums_heritage', 'nature_landscape'],
        'travel_style': 'cultural'
    }),
    ("Krakow (utf8)", {
        'location': {'city': 'Krak\u00f3w', 'country': 'Poland', 'region_type': 'city'},
        'group': {'type': 'couples', 'size': 2, 'crowd_tolerance': 2},
        'trip_length': {'days': 2, 'start_date': '2026-06-07'},
        'daily_time_window': {'start': '09:00', 'end': '19:00'},
        'budget': {'level': 2},
        'transport_modes': ['walk'],
        'preferences': ['museums_heritage', 'nature_landscape'],
        'travel_style': 'cultural'
    }),
    ("Wroclaw (ascii)", {
        'location': {'city': 'Wroclaw', 'country': 'Poland', 'region_type': 'city'},
        'group': {'type': 'couples', 'size': 2, 'crowd_tolerance': 2},
        'trip_length': {'days': 2, 'start_date': '2026-06-07'},
        'daily_time_window': {'start': '09:00', 'end': '19:00'},
        'budget': {'level': 2},
        'transport_modes': ['walk'],
        'preferences': ['museums_heritage'],
        'travel_style': 'cultural'
    }),
    ("Gdansk (ascii)", {
        'location': {'city': 'Gdansk', 'country': 'Poland', 'region_type': 'city'},
        'group': {'type': 'couples', 'size': 2, 'crowd_tolerance': 2},
        'trip_length': {'days': 2, 'start_date': '2026-06-07'},
        'daily_time_window': {'start': '09:00', 'end': '19:00'},
        'budget': {'level': 2},
        'transport_modes': ['walk'],
        'preferences': ['museums_heritage'],
        'travel_style': 'cultural'
    }),
]

for name, body in tests:
    print(f"\n=== {name} ===")
    try:
        r = requests.post('http://127.0.0.1:8001/plan/preview', json=body, headers=headers, timeout=60)
        print(f"HTTP {r.status_code}")
        if r.ok:
            plan = r.json()
            days = plan.get('days', [])
            print(f"Days: {len(days)}")
            for d in days:
                items = d.get('items', [])
                attrs = [i for i in items if i.get('type') == 'attraction']
                print(f"  Day {d.get('day')}: {len(items)} items, {len(attrs)} attractions")
                for a in attrs:
                    print(f"    - {a.get('name')}")
        else:
            print(r.text[:800])
    except Exception as e:
        print(f"EXCEPTION: {type(e).__name__}: {e}")

print("\nDone.")

# --- Part 2: direct data check ---
print("\n\n=== DIRECT DATA CHECK ===")
from app.infrastructure.repositories.load_multi_city import load_multi_city_poi

for city in ['Krak\u00f3w', 'Krakow', 'Wroc\u0142aw', 'Gda\u0144sk']:
    pois = load_multi_city_poi('data/multi_city_attractions.xlsx', [city])
    cities_found = set(p.get('city', '') for p in pois)
    print(f"  '{city}' -> {len(pois)} POIs, cities_in_data: {cities_found}")

from app.domain.planner.engine import is_open, is_core_poi
from datetime import datetime

pois = load_multi_city_poi('data/multi_city_attractions.xlsx', ['Krakow'])
print(f'Total: {len(pois)} POIs for Krakow (ASCII)')

pois2 = load_multi_city_poi('data/multi_city_attractions.xlsx', ['Kraków'])
print(f'Total: {len(pois2)} POIs for Kraków (Polish)')

# Use whichever has results
all_pois = pois if len(pois) > 0 else pois2

if not all_pois:
    print("NO POIS FOUND - check city name in Excel")
    import openpyxl
    wb = openpyxl.load_workbook('data/multi_city_attractions.xlsx')
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    print(f"Headers: {headers}")
    # print first few city values
    for row in ws.iter_rows(min_row=2, max_row=5, values_only=True):
        print(dict(zip(headers, row)))
    sys.exit()

print(f"\nSample POI keys: {list(all_pois[0].keys())}")
print(f"\nFirst 3 POIs:")
for p in all_pois[:3]:
    print(f"  name={p.get('name')}, priority_level={p.get('priority_level')}, duration_min={p.get('duration_min')}, time_min={p.get('time_min')}")
    oh = p.get('opening_hours')
    print(f"  opening_hours type={type(oh).__name__}, val={str(oh)[:100]}")

# Check is_core_poi for all
core_count = sum(1 for p in all_pois if is_core_poi(p))
print(f"\nCore POIs: {core_count}/{len(all_pois)}")

# Check is_open
context = {'date': datetime(2025, 6, 20)}
open_results = []
for p in all_pois:
    result = is_open(p, 600, 90, 'summer', context)
    open_results.append(result)
open_count = sum(1 for r in open_results if r)
print(f"\nOpen at 10:00: {open_count}/{len(all_pois)}")
if open_count < len(all_pois):
    # Show a closed POI
    for i, (p, r) in enumerate(zip(all_pois, open_results)):
        if not r:
            print(f"  CLOSED: {p.get('name')}, OH={p.get('opening_hours')}")
            if i > 5:
                break

# Check priority field
has_priority = sum(1 for p in all_pois if p.get('priority') is not None)
print(f"\nHas 'priority' field: {has_priority}/{len(all_pois)}")
has_type_of_attr = sum(1 for p in all_pois if p.get('type_of_attraction'))
print(f"Has 'type_of_attraction': {has_type_of_attr}/{len(all_pois)}")
has_time_min = sum(1 for p in all_pois if p.get('time_min'))
print(f"Has 'time_min': {has_time_min}/{len(all_pois)}")
has_Name = sum(1 for p in all_pois if p.get('Name'))
print(f"Has 'Name' (uppercase): {has_Name}/{len(all_pois)}")
