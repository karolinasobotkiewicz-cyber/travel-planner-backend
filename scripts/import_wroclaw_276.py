# -*- coding: utf-8 -*-
"""FIX #276 — Wrocław-only import from client Excel (attractions + restaurants).

Does NOT touch other cities. Does NOT delete rows. Does NOT steal Ząbkowice
from Hub=Kłodzko (those stay shared via city_excel + satellite load).
"""
from __future__ import annotations

import math
import os
import shutil
import sys
import uuid
from datetime import datetime
from pathlib import Path

import pandas as pd
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
PARENT = ROOT.parent
NEW_ATTR = Path(r"C:\Users\matte\Downloads\Planer - miasta atrakcje.xlsx")
NEW_REST = Path(r"C:\Users\matte\Downloads\Planer - restauracje.xlsx")
MULTI = ROOT / "data" / "multi_city_attractions.xlsx"
SHEET = "All Cities"

# Client rows whose City/GPS is known-wrong — never overwrite identity/coords.
_SKIP_IDENTITY_MARKERS = (
    "ogród botaniczny uniwersytetu wrocławskiego",
    "ogrod botaniczny uniwersytetu wroclawskiego",
    "muzeum świat iluzji we wrocławiu",
    "muzeum swiat iluzji we wroclawiu",
    "fontanna multimedialna",
)

_UPDATE_COLS = (
    "recommended_time_of_day",
    "Type of attraction",
    "Activity_style",
    "Tags",
    "priority_level",
    "Seasonality of attractions",
    "weather_dependency",
    "Must see score",
    "Pro_tip",
    "Description_short",
    "Description_long",
    "Why visit",
    "kids_only",
    "time_min",
    "time_max",
    "Opening hours",
    "opening_hours_seasonal",
    "Zone",
    "popularity_score",
    "Intensity",
    "Space",
    "crowd_level",
    "image_key",
    "Price",
    "ticket_normal",
    "ticket_reduced",
    "Address",
    "parking_name",
    "parking_address",
    "parking_lat",
    "parking_lng",
    "parking_type",
    "parking_walk_time_min",
)

_EXISTING_ZABKOWICE = {
    "krzywa wieża w ząbkowicach śląskich",
    "krzywa wieza w zabkowicach slaskich",
    "zamek w ząbkowicach śląskich",
    "zamek w zabkowicach slaskich",
    "rynek w ząbkowicach śląskich",
    "rynek w zabkowicach slaskich",
    "laboratorium frankensteina",
    "laboratorium dr. frankensteina",
    "laboratorium dr frankensteina",
}


def _norm(s) -> str:
    return str(s or "").strip()


def _norm_name(s) -> str:
    import unicodedata
    t = unicodedata.normalize("NFKD", _norm(s).lower())
    return "".join(c for c in t if not unicodedata.combining(c))


def _is_nan(v) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return True
    s = str(v).strip().lower()
    return s in ("", "nan", "none", "-")


def _cell(v):
    return None if _is_nan(v) else v


def _skip_identity(name: str) -> bool:
    n = _norm_name(name)
    return any(m in n for m in _SKIP_IDENTITY_MARKERS)


def _is_existing_zabkowice(name: str) -> bool:
    return _norm_name(name) in _EXISTING_ZABKOWICE or any(
        m in _norm_name(name) for m in _EXISTING_ZABKOWICE
    )


def _haversine(lat1, lng1, lat2, lng2) -> float:
    r = 6371.0
    p1, p2 = math.radians(float(lat1)), math.radians(float(lat2))
    dphi = math.radians(float(lat2) - float(lat1))
    dl = math.radians(float(lng2) - float(lng1))
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * r * math.asin(math.sqrt(min(1.0, a)))


def import_attractions() -> dict:
    df_new = pd.read_excel(NEW_ATTR, sheet_name="Wrocław")
    df_new = df_new.rename(columns={c: str(c).strip() for c in df_new.columns})
    df_cur = pd.read_excel(MULTI, sheet_name=SHEET)

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = ROOT / "data" / f"multi_city_attractions_backup_{stamp}.xlsx"
    shutil.copy2(MULTI, backup)

    wb = openpyxl.load_workbook(MULTI)
    ws = wb[SHEET]
    header = [cell.value for cell in ws[1]]
    col = {name: i + 1 for i, name in enumerate(header) if name}

    # index existing rows by normalized name
    existing = {}
    for r in range(2, ws.max_row + 1):
        name = _norm(ws.cell(r, col["Name"]).value)
        if not name:
            continue
        existing.setdefault(_norm_name(name), []).append(r)

    added, updated, skipped = [], [], []
    for _, row in df_new.iterrows():
        name = _norm(row.get("Name"))
        if not name:
            continue
        key = _norm_name(name)
        # Laboratorium Frankensteina == existing "Laboratorium dr. Frankensteina"
        if "laboratorium frankensteina" in key and "dr" not in key:
            match_key = "laboratorium dr. frankensteina"
        else:
            match_key = key
        hits = existing.get(match_key) or existing.get(key) or []

        if hits:
            # Prefer a Wrocław-hub / Wrocław-city row when several exist
            row_i = hits[0]
            if len(hits) > 1:
                for h in hits:
                    city = _norm(ws.cell(h, col.get("City", 1)).value)
                    hub = _norm(ws.cell(h, col["Hub"]).value) if "Hub" in col else ""
                    if "wroc" in city.lower() or "wroc" in hub.lower():
                        row_i = h
                        break
            # Never rewrite identity on known-bad client rows
            if _skip_identity(name):
                # still allow tag/type/tod updates if coords in Excel current are Wrocław
                cur_lat = ws.cell(row_i, col["Lat"]).value
                new_lat = row.get("Lat")
                try:
                    if cur_lat and new_lat and _haversine(cur_lat, ws.cell(row_i, col["Lng"]).value, new_lat, row.get("Lng")) > 20:
                        skipped.append(f"skip-coords {name}")
                        continue
                except Exception:
                    skipped.append(f"skip-bad {name}")
                    continue
            # Do not change Hub/City on shared Ząbkowice (Kotlina)
            change_hub = not _is_existing_zabkowice(name)
            for cname in _UPDATE_COLS:
                if cname not in col or cname not in df_new.columns:
                    continue
                val = row.get(cname)
                if _is_nan(val):
                    continue
                ws.cell(row_i, col[cname]).value = _cell(val)
            if change_hub and "Hub" in col:
                hub_now = _norm(ws.cell(row_i, col["Hub"]).value)
                if not hub_now:
                    ws.cell(row_i, col["Hub"]).value = "Wrocław"
            # City: only if client city is a real satellite / Wrocław, not Warsaw/Kraków
            new_city = _norm(row.get("City"))
            if new_city and new_city.lower() not in ("warszawa", "warsaw", "kraków", "krakow"):
                if not _skip_identity(name) and not _is_existing_zabkowice(name):
                    ws.cell(row_i, col["City"]).value = new_city
            updated.append(name)
            continue

        # New row
        if _is_existing_zabkowice(name):
            skipped.append(f"already-shared {name}")
            continue
        last = ws.max_row + 1
        for cname, cidx in col.items():
            if cname == "Hub":
                ws.cell(last, cidx).value = "Wrocław"
                continue
            val = row.get(cname) if cname in df_new.columns else None
            ws.cell(last, cidx).value = _cell(val)
        existing.setdefault(key, []).append(last)
        added.append(name)

    wb.save(MULTI)
    return {
        "backup": str(backup),
        "added": added,
        "updated": updated,
        "skipped": skipped,
        "cur_before": len(df_cur),
    }


def upsert_restaurants() -> dict:
    sys.path.insert(0, str(ROOT))
    from app.infrastructure.database import RestaurantDB
    from app.infrastructure.database.connection import SessionLocal
    from scripts.data_mappings import parse_meal_type, clean_string, safe_float, safe_int
    from scripts.load_restaurants import (
        generate_uuid, parse_tags, parse_target_group, parse_bool,
    )

    df = pd.read_excel(NEW_REST, sheet_name="Wrocław")
    session = SessionLocal()
    inserted, merged, skipped = [], [], []
    for _, row in df.iterrows():
        name = clean_string(row.get("Name"))
        if not name:
            continue
        meal_type = parse_meal_type(str(row.get("meal_type", "")))
        if meal_type is None:
            skipped.append(f"coffee {name}")
            continue
        city = clean_string(row.get("City")) or "Wrocław"
        lat = safe_float(row.get("Lat"))
        lng = safe_float(row.get("Lng"))
        rid = generate_uuid(name, city, lat)
        obj = RestaurantDB(
            id=rid,
            name=name,
            city=city,
            meal_type=meal_type,
            cuisine_type=clean_string(row.get("cuisine_type", "")) or None,
            place_type=clean_string(row.get("place_type", "")) or None,
            lat=lat,
            lng=lng,
            address=clean_string(row.get("Address", "")),
            visit_duration_min=safe_int(row.get("visit_duration_min", 60), default=60),
            visit_duration_max=safe_int(row.get("visit_duration_max", 90), default=90),
            price_level=safe_int(row.get("price_level", 2), default=2),
            avg_meal_cost=50,
            children_friendly=True,
            target_group=parse_target_group(str(row.get("recommended_for", ""))),
            pro_tip=clean_string(row.get("pro_tip", "")) or None,
            reservations_required=parse_bool(row.get("reservation_recommended", False)),
            parking_name=clean_string(row.get("parking_name", "")) or None,
            parking_lat=safe_float(row.get("parking_lat", 0.0)) or None,
            parking_lng=safe_float(row.get("parking_lng", 0.0)) or None,
            parking_walk_time_min=safe_int(row.get("parking_walk_time_min", 0)) or 0,
            image_key=clean_string(row.get("image_key", "")) or None,
            tags=parse_tags(str(row.get("Tags", ""))) or None,
        )
        existed = session.get(RestaurantDB, rid) is not None
        session.merge(obj)
        (merged if existed else inserted).append(f"{name} ({city})")
    session.commit()
    session.close()
    return {"inserted": inserted, "merged": merged, "skipped": skipped}


def main():
    print("=== FIX #276 Wrocław import ===")
    # provenance copies (parent planer_update) — other cities come tomorrow
    dest_dir = PARENT / "planer_update"
    dest_dir.mkdir(exist_ok=True)
    shutil.copy2(NEW_ATTR, dest_dir / "Planer - miasta atrakcje.xlsx")
    shutil.copy2(NEW_REST, dest_dir / "Planer - restauracje.xlsx")

    attr = import_attractions()
    print(f"backup: {attr['backup']}")
    print(f"ADDED ({len(attr['added'])}):")
    for n in attr["added"]:
        print(f"  + {n}")
    print(f"UPDATED ({len(attr['updated'])})")
    print(f"SKIPPED ({len(attr['skipped'])}): {attr['skipped']}")

    rest = upsert_restaurants()
    print(f"REST inserted ({len(rest['inserted'])}): {rest['inserted']}")
    print(f"REST merged ({len(rest['merged'])})")
    print(f"REST skipped ({len(rest['skipped'])}): {rest['skipped']}")


if __name__ == "__main__":
    main()
