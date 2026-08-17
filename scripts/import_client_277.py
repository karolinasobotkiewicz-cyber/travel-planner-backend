# -*- coding: utf-8 -*-
"""FIX #277 — full client Excel (all cities). Add new rows, update exact matches.

Does not delete. Does not steal Hub. Skips known-bad identity/coords and
Duszniki/Kudowa 'midday.afternoon' typos. Fuzzy-renames are NOT added.
"""
from __future__ import annotations

import math
import shutil
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).parent.parent
PARENT = ROOT.parent
NEW_ATTR = Path(r"C:\Users\matte\Downloads\Planer - miasta atrakcje.xlsx")
NEW_REST = Path(r"C:\Users\matte\Downloads\Planer - restauracje.xlsx")
MULTI = ROOT / "data" / "multi_city_attractions.xlsx"
ZAK = ROOT / "data" / "zakopane.xlsx"
SHEET = "All Cities"

_SKIP_IDENTITY = (
    "ogród botaniczny uniwersytetu wrocławskiego",
    "ogrod botaniczny uniwersytetu wroclawskiego",
    "muzeum świat iluzji we wrocławiu",
    "muzeum swiat iluzji we wroclawiu",
    "fontanna multimedialna",
)

_UPDATE_COLS = (
    "recommended_time_of_day", "Type of attraction", "Activity_style", "Tags",
    "priority_level", "Seasonality of attractions", "weather_dependency",
    "Must see score", "Pro_tip", "Description_short", "Description_long",
    "Why visit", "kids_only", "time_min", "time_max", "Opening hours",
    "opening_hours_seasonal", "Zone", "popularity_score", "Intensity", "Space",
    "crowd_level", "image_key", "Price", "ticket_normal", "ticket_reduced",
    "Address", "parking_name", "parking_address", "parking_lat", "parking_lng",
    "parking_type", "parking_walk_time_min",
)


def fold(s) -> str:
    t = unicodedata.normalize("NFKD", str(s or "").strip().lower())
    t = "".join(c for c in t if not unicodedata.combining(c))
    return t.replace("ł", "l").replace("\u0142", "l")


def _is_nan(v) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return True
    return str(v).strip().lower() in ("", "nan", "none", "-")


def _cell(v):
    return None if _is_nan(v) else v


def _hav(lat1, lng1, lat2, lng2) -> float:
    try:
        r = 6371.0
        p1, p2 = math.radians(float(lat1)), math.radians(float(lat2))
        dphi = math.radians(float(lat2) - float(lat1))
        dl = math.radians(float(lng2) - float(lng1))
        a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
        return 2 * r * math.asin(math.sqrt(min(1.0, a)))
    except Exception:
        return 999.0


def _skip_identity(name: str) -> bool:
    n = fold(name)
    return any(m in n for m in (fold(x) for x in _SKIP_IDENTITY))


def _tod_is_typo(val) -> bool:
    s = str(val or "")
    return "midday.afternoon" in s.replace(" ", "")


def _norm_tag(val) -> str:
    return " ".join(str(val or "").replace("\n", ",").replace(";", ",").split()).lower()


def import_attractions() -> dict:
    frames = []
    xl = pd.ExcelFile(NEW_ATTR)
    for s in xl.sheet_names:
        df = pd.read_excel(NEW_ATTR, sheet_name=s)
        df = df.rename(columns={c: str(c).strip() for c in df.columns})
        df["_sheet"] = s
        frames.append(df)
    df_new = pd.concat(frames, ignore_index=True)

    zak_names = set()
    if ZAK.exists():
        zdf = pd.read_excel(ZAK)
        zdf = zdf.rename(columns={c: str(c).strip() for c in zdf.columns})
        zak_names = {fold(n) for n in zdf.get("Name", pd.Series(dtype=str)).dropna()}

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = ROOT / "data" / f"multi_city_attractions_backup_{stamp}.xlsx"
    shutil.copy2(MULTI, backup)

    wb = openpyxl.load_workbook(MULTI)
    ws = wb[SHEET]
    header = [cell.value for cell in ws[1]]
    col = {name: i + 1 for i, name in enumerate(header) if name}

    existing = []
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, col["Name"]).value or "").strip()
        if not name:
            continue
        existing.append({
            "row": r,
            "name": name,
            "fold": fold(name),
            "city": fold(ws.cell(r, col["City"]).value),
            "hub": fold(ws.cell(r, col["Hub"]).value) if "Hub" in col else "",
            "lat": ws.cell(r, col["Lat"]).value,
            "lng": ws.cell(r, col["Lng"]).value,
        })

    added, updated, skipped = [], [], []

    def _distinct(s: str) -> set:
        stop = {
            "muzeum", "park", "rynek", "centrum", "w", "na", "do", "im", "sw",
            "the", "and", "przy", "oraz", "dla", "pod",
        }
        return {t for t in fold(s).replace("-", " ").split() if len(t) > 3 and t not in stop}

    def find_row(name, city, lat, lng, sheet):
        nf, cf = fold(name), fold(city)
        if not nf or nf in ("nan", "none"):
            return None, None
        for p in existing:
            if p["fold"] != nf:
                continue
            if cf and (cf == p["city"] or cf == p["hub"] or cf == fold(sheet)):
                return p, "exact-city"
            if _hav(lat, lng, p["lat"], p["lng"]) < 0.5:
                return p, "exact-near"
        for p in existing:
            if "frankenstein" in nf and "frankenstein" in p["fold"]:
                if _hav(lat, lng, p["lat"], p["lng"]) < 0.4:
                    return p, "frankenstein"
            if nf not in p["fold"] and p["fold"] not in nf:
                ds, dp = _distinct(name), _distinct(p["name"])
                if ds and dp and (ds <= dp or dp <= ds) and _hav(lat, lng, p["lat"], p["lng"]) < 0.3:
                    return p, "tokens-near"
                continue
            if min(len(nf), len(p["fold"])) < 12:
                continue
            if _hav(lat, lng, p["lat"], p["lng"]) < 0.45:
                return p, "rename-near"
        return None, None

    for _, row in df_new.iterrows():
        name = str(row.get("Name") or "").strip()
        if not name or fold(name) in ("nan", "none"):
            continue
        sheet = str(row.get("_sheet") or "")
        city = str(row.get("City") or "").strip()
        if city.lower() in ("nan", "none"):
            city = ""
        lat, lng = row.get("Lat"), row.get("Lng")
        hit, why = find_row(name, city, lat, lng, sheet)

        if hit:
            if _skip_identity(name):
                skipped.append(f"identity {name}")
                continue
            row_i = hit["row"]
            for cname in _UPDATE_COLS:
                if cname not in col or cname not in df_new.columns:
                    continue
                val = row.get(cname)
                if _is_nan(val):
                    continue
                if cname == "recommended_time_of_day" and _tod_is_typo(val):
                    continue
                if cname == "Tags" and _norm_tag(val) == _norm_tag(ws.cell(row_i, col[cname]).value):
                    continue
                ws.cell(row_i, col[cname]).value = _cell(val)
            if city and city.lower() not in ("warsaw", "warszawa", "kraków", "krakow"):
                # never blank Hub; never overwrite a different cluster hub
                pass
            updated.append(f"{name} ({why})")
            continue

        if sheet == "Zakopane" and fold(name) in zak_names:
            skipped.append(f"zak-xlsx {name}")
            continue
        if _skip_identity(name):
            skipped.append(f"skip-new-identity {name}")
            continue

        last = ws.max_row + 1
        for cname, cidx in col.items():
            if cname == "Hub":
                ws.cell(last, cidx).value = sheet
                continue
            val = row.get(cname) if cname in df_new.columns else None
            ws.cell(last, cidx).value = _cell(val)
        if not city and "City" in col:
            ws.cell(last, col["City"]).value = sheet
        existing.append({
            "row": last, "name": name, "fold": fold(name),
            "city": fold(city or sheet), "hub": fold(sheet),
            "lat": lat, "lng": lng,
        })
        added.append(f"{name} [{sheet}/{city or sheet}]")

    wb.save(MULTI)
    return {"backup": str(backup), "added": added, "updated": len(updated), "skipped": skipped}


def upsert_restaurants() -> dict:
    sys.path.insert(0, str(ROOT))
    from app.infrastructure.database import RestaurantDB
    from app.infrastructure.database.connection import SessionLocal
    from scripts.data_mappings import parse_meal_type, clean_string, safe_float, safe_int
    from scripts.load_restaurants import (
        generate_uuid, parse_tags, parse_target_group, parse_bool,
    )

    xl = pd.ExcelFile(NEW_REST)
    session = SessionLocal()
    inserted, merged, skipped = [], [], []
    for sheet in xl.sheet_names:
        df = pd.read_excel(NEW_REST, sheet_name=sheet)
        for _, row in df.iterrows():
            name = clean_string(row.get("Name"))
            if not name:
                continue
            meal_type = parse_meal_type(str(row.get("meal_type", "")))
            if meal_type is None:
                skipped.append(f"coffee {name}")
                continue
            city = clean_string(row.get("City")) or sheet
            lat = safe_float(row.get("Lat"))
            lng = safe_float(row.get("Lng"))
            rid = generate_uuid(name, city, lat)
            obj = RestaurantDB(
                id=rid,
                name=name,
                city=city,
                meal_type=meal_type,
                cuisine_type=clean_string(row.get("cuisine_type", "")) or None,
                place_type=clean_string(row.get("place_type", "")) or None,
                lat=lat,
                lng=lng,
                address=clean_string(row.get("Address", "")),
                visit_duration_min=safe_int(row.get("visit_duration_min", 60), default=60),
                visit_duration_max=safe_int(row.get("visit_duration_max", 90), default=90),
                price_level=safe_int(row.get("price_level", 2), default=2),
                avg_meal_cost=50,
                children_friendly=True,
                target_group=parse_target_group(str(row.get("recommended_for", ""))),
                pro_tip=clean_string(row.get("pro_tip", "")) or None,
                reservations_required=parse_bool(row.get("reservation_recommended", False)),
                parking_name=clean_string(row.get("parking_name", "")) or None,
                parking_lat=safe_float(row.get("parking_lat", 0.0)) or None,
                parking_lng=safe_float(row.get("parking_lng", 0.0)) or None,
                parking_walk_time_min=safe_int(row.get("parking_walk_time_min", 0)) or 0,
                image_key=clean_string(row.get("image_key", "")) or None,
                tags=parse_tags(str(row.get("Tags", ""))) or None,
            )
            existed = session.get(RestaurantDB, rid) is not None
            session.merge(obj)
            (merged if existed else inserted).append(f"{name} ({city})")
    session.commit()
    session.close()
    return {"inserted": inserted, "merged": len(merged), "skipped": len(skipped)}


def main():
    print("=== FIX #277 client Excel import ===")
    dest = PARENT / "planer_update"
    dest.mkdir(exist_ok=True)
    shutil.copy2(NEW_ATTR, dest / "Planer - miasta atrakcje.xlsx")
    shutil.copy2(NEW_REST, dest / "Planer - restauracje.xlsx")
    attr = import_attractions()
    print(f"backup: {attr['backup']}")
    print(f"ADDED ({len(attr['added'])}):")
    for n in attr["added"]:
        print(f"  + {n}")
    print(f"UPDATED {attr['updated']} exact/near matches")
    print(f"SKIPPED ({len(attr['skipped'])}): {attr['skipped'][:20]}")
    rest = upsert_restaurants()
    print(f"REST inserted ({len(rest['inserted'])}):")
    for n in rest["inserted"]:
        print(f"  + {n}")
    print(f"REST merged {rest['merged']}, skipped coffee {rest['skipped']}")


if __name__ == "__main__":
    main()
