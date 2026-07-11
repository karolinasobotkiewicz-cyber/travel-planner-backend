"""
FIX #235 — sync Wrocław POI updates from client file to multi_city_attractions.xlsx.

Actions:
1. Update Arboretum Wojsławice opening_hours_seasonal from planer8
2. Remove Bungee Wrocław (client removed from their database)
"""
from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

PROJECT = Path(__file__).parent.parent
WORKSPACE = PROJECT.parent
MC_PATH = PROJECT / "data" / "multi_city_attractions.xlsx"
PLANER8 = WORKSPACE / "planer_update" / "Planer - miasta atrakcje8.xlsx"


def main() -> None:
    if not MC_PATH.exists():
        raise FileNotFoundError(MC_PATH)
    if not PLANER8.exists():
        raise FileNotFoundError(PLANER8)

    backup = MC_PATH.with_name(
        f"multi_city_attractions_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    shutil.copy2(MC_PATH, backup)
    print(f"Backup: {backup}")

    p8 = pd.read_excel(PLANER8, sheet_name="Wrocław")
    arb_p8 = p8[p8["Name"].fillna("").str.contains("Wojsławice", case=False, na=False)]
    if arb_p8.empty:
        raise ValueError("Arboretum Wojsławice not found in planer8 Wrocław sheet")
    new_oh = arb_p8.iloc[0].get("opening_hours_seasonal")
    new_season = arb_p8.iloc[0].get("Seasonality of attractions")

    wb = openpyxl.load_workbook(MC_PATH)
    ws = wb["All Cities"]
    header = {cell.value: cell.column for cell in ws[1]}
    col_name = header["Name"]
    col_oh = header.get("opening_hours_seasonal")
    col_season = header.get("Seasonality of attractions")

    updated = 0
    deleted = 0
    rows_to_delete = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        name = str(row[col_name - 1].value or "").strip()
        if "bungee" in name.lower() and "wroc" in name.lower():
            rows_to_delete.append(row_idx)
            continue
        if "wojsławice" in name.lower() or "wojslawice" in name.lower():
            if col_oh and new_oh is not None and str(new_oh).strip():
                row[col_oh - 1].value = str(new_oh)
            if col_season and new_season is not None and str(new_season).strip():
                row[col_season - 1].value = str(new_season)
            updated += 1
            print(f"Updated: {name}")

    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx, 1)
        deleted += 1
        print("Deleted: Bungee Wrocław")

    wb.save(MC_PATH)
    print(f"Done: updated={updated}, deleted={deleted}, rows now={ws.max_row - 1}")


if __name__ == "__main__":
    main()
